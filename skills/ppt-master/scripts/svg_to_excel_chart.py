#!/usr/bin/env python3
"""
PPT Master - SVG Chart to Excel Chart Tool

Extract numeric chart data from Executor SVG pages and write real Excel chart
objects (bar / line / pie) with openpyxl. Also exposes extraction helpers for
PPTX export and quality checking.

Usage:
    python3 scripts/svg_to_excel_chart.py convert --svg PATH --output PATH [--chart-type auto|bar|line|pie]

Examples:
    python3 scripts/svg_to_excel_chart.py convert \\
        --svg templates/charts/bar_chart.svg \\
        --output /tmp/bar_chart.xlsx

Dependencies:
    openpyxl>=3.1.0 (optional — required for convert / XLSX output)
"""

from __future__ import annotations

import argparse
import html
import json
import re
import sys
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

_SCRIPTS_DIR = Path(__file__).resolve().parent
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
except ImportError:
    Workbook = None  # type: ignore[misc, assignment]
    BarChart = LineChart = PieChart = Reference = None  # type: ignore[misc, assignment]

SVG_NS = 'http://www.w3.org/2000/svg'
PLOT_AREA_RE = re.compile(
    r'chart-plot-area:\s*'
    r'(?:(?P<xmin>\d+(?:\.\d+)?)\s*,\s*(?P<ymin>\d+(?:\.\d+)?)\s*,\s*'
    r'(?P<xmax>\d+(?:\.\d+)?)\s*,\s*(?P<ymax>\d+(?:\.\d+)?)'
    r'|(?P<radial>\w+)\s*\|\s*center:\s*(?P<cx>\d+(?:\.\d+)?)\s*,\s*(?P<cy>\d+(?:\.\d+)?)'
    r'(?:\s*\|\s*radius:\s*(?P<radius>\d+(?:\.\d+)?))?'
    r'(?:\s*\|\s*outer-radius:\s*(?P<outer>\d+(?:\.\d+)?))?'
    r'(?:\s*\|\s*inner-radius:\s*(?P<inner>\d+(?:\.\d+)?))?)',
    re.IGNORECASE,
)
BAR_GROUP_RE = re.compile(r'^bar[-_]?\d+$', re.IGNORECASE)
NUMERIC_RE = re.compile(r'-?\d+(?:\.\d+)?')
PERCENT_RE = re.compile(r'(\d+(?:\.\d+)?)\s*%')


def _local_tag(tag: str) -> str:
    return tag.split('}', 1)[-1] if '}' in tag else tag


def _parse_float(value: str | None, default: float | None = None) -> float | None:
    if value is None:
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _strip_text(raw: str) -> str:
    text = html.unescape(raw or '')
    text = re.sub(r'<[^>]+>', ' ', text)
    return re.sub(r'\s+', ' ', text).strip()


def _element_text(elem: ET.Element) -> str:
    parts = [_strip_text(ET.tostring(child, encoding='unicode', method='text'))
             for child in list(elem)]
    if not parts:
        parts = [_strip_text(elem.text or '')]
    return ' '.join(p for p in parts if p).strip()


def _parse_numeric(text: str) -> float | None:
    cleaned = text.replace(',', '').replace('$', '').replace('B', '').replace('M', '')
    match = NUMERIC_RE.search(cleaned)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def _find_chart_roots(root: ET.Element) -> list[ET.Element]:
    """Return chart container elements (id contains 'chart' or class='chart')."""
    charts: list[ET.Element] = []
    for elem in root.iter():
        elem_id = (elem.get('id') or '').lower()
        elem_class = (elem.get('class') or '').lower()
        if 'chart' in elem_id or elem_class == 'chart':
            charts.append(elem)
    if not charts:
        chart_area = root.find(f".//{{{SVG_NS}}}g[@id='chartArea']")
        if chart_area is not None:
            charts.append(chart_area)
        for elem in root.iter():
            if _local_tag(elem.tag) == 'g' and 'pie' in (elem.get('id') or '').lower():
                if elem not in charts:
                    charts.append(elem)
                break
    # Prefer outermost containers (chartArea / pieChart) over nested bar-* groups.
    parent_map = root_parent_map(root)
    filtered: list[ET.Element] = []
    for candidate in charts:
        if any(
            candidate is not other and _is_descendant(other, candidate, parent_map)
            for other in charts
        ):
            continue
        filtered.append(candidate)
    return filtered or charts


def _is_descendant(ancestor: ET.Element, node: ET.Element, parent_map: dict[ET.Element, ET.Element]) -> bool:
    current = node
    while current is not None:
        if current is ancestor:
            return True
        current = parent_map.get(current)
    return False


def root_parent_map(root: ET.Element) -> dict[ET.Element, ET.Element]:
    mapping: dict[ET.Element, ET.Element] = {}
    for parent in root.iter():
        for child in list(parent):
            mapping[child] = parent
    return mapping


def parse_plot_area_marker(svg_text: str) -> dict[str, Any]:
    match = PLOT_AREA_RE.search(svg_text)
    if not match:
        return {}
    if match.group('xmin'):
        return {
            'kind': 'rect',
            'x_min': float(match.group('xmin')),
            'y_min': float(match.group('ymin')),
            'x_max': float(match.group('xmax')),
            'y_max': float(match.group('ymax')),
        }
    return {
        'kind': match.group('radial') or 'pie',
        'center_x': _parse_float(match.group('cx')),
        'center_y': _parse_float(match.group('cy')),
        'radius': _parse_float(match.group('radius') or match.group('outer')),
        'inner_radius': _parse_float(match.group('inner')),
    }


def detect_chart_type(root: ET.Element, svg_text: str, chart_root: ET.Element | None = None) -> str:
    scope = chart_root if chart_root is not None else root
    scope_text = ET.tostring(scope, encoding='unicode')

    if re.search(r'id=["\']pie', scope_text, re.IGNORECASE) or ' pie ' in svg_text.lower():
        if scope.findall(f".//{{{SVG_NS}}}path") and PLOT_AREA_RE.search(svg_text):
            radial = parse_plot_area_marker(svg_text)
            if radial.get('kind') in ('pie', 'donut'):
                return 'pie'

    bar_groups = [
        elem for elem in scope.iter()
        if _local_tag(elem.tag) == 'g' and BAR_GROUP_RE.match(elem.get('id') or '')
    ]
    vertical_bars = [
        rect for rect in scope.iter()
        if _local_tag(rect.tag) == 'rect'
        and _parse_float(rect.get('width'), 0) is not None
        and _parse_float(rect.get('height'), 0) is not None
        and (_parse_float(rect.get('width'), 0) or 0) < (_parse_float(rect.get('height'), 0) or 0) * 3
    ]
    if bar_groups or len(vertical_bars) >= 2:
        return 'bar'

    polylines = scope.findall(f".//{{{SVG_NS}}}polyline")
    if polylines:
        return 'line'

    paths = scope.findall(f".//{{{SVG_NS}}}path")
    if paths and parse_plot_area_marker(svg_text).get('kind') in ('pie', 'donut'):
        return 'pie'

    circles = scope.findall(f".//{{{SVG_NS}}}circle")
    if circles and polylines:
        return 'line'

    return 'bar'


def extract_bar_data(root: ET.Element, chart_root: ET.Element | None = None) -> dict[str, Any]:
    scope = chart_root if chart_root is not None else root
    parent_map = root_parent_map(root)
    entries: list[dict[str, Any]] = []

    bar_groups = [
        elem for elem in scope.iter()
        if _local_tag(elem.tag) == 'g' and BAR_GROUP_RE.match(elem.get('id') or '')
    ]

    if bar_groups:
        for group in sorted(bar_groups, key=lambda g: g.get('id', '')):
            rect = next((c for c in group if _local_tag(c.tag) == 'rect'), None)
            if rect is None:
                continue
            texts = [c for c in group if _local_tag(c.tag) == 'text']
            value_text = ''
            category = ''
            rect_y = _parse_float(rect.get('y'), 0) or 0
            for text_elem in texts:
                content = _element_text(text_elem)
                y = _parse_float(text_elem.get('y'), 0) or 0
                if y < rect_y and _parse_numeric(content) is not None:
                    value_text = content
                elif y > rect_y:
                    category = content or category
            value = _parse_numeric(value_text)
            if value is None:
                height = _parse_float(rect.get('height'))
                if height is not None:
                    value = height
            if category or value is not None:
                cx = (_parse_float(rect.get('x'), 0) or 0) + (_parse_float(rect.get('width'), 0) or 0) / 2
                entries.append({
                    'category': category or f'Item {len(entries) + 1}',
                    'value': value if value is not None else 0.0,
                    'x': cx,
                })
    else:
        rects = [
            rect for rect in scope.iter()
            if _local_tag(rect.tag) == 'rect'
            and (_parse_float(rect.get('height'), 0) or 0) > 20
            and (_parse_float(rect.get('width'), 0) or 0) > 10
        ]
        for rect in rects:
            width = _parse_float(rect.get('width'), 0) or 0
            height = _parse_float(rect.get('height'), 0) or 0
            if height <= width:
                continue
            group = rect
            while group in parent_map and _local_tag(parent_map[group].tag) == 'g':
                group = parent_map[group]
            texts = [c for c in group.iter() if _local_tag(c.tag) == 'text']
            category = ''
            value = None
            rect_y = _parse_float(rect.get('y'), 0) or 0
            for text_elem in texts:
                content = _element_text(text_elem)
                y = _parse_float(text_elem.get('y'), 0) or 0
                if y < rect_y:
                    value = _parse_numeric(content) if _parse_numeric(content) is not None else value
                elif y > rect_y:
                    category = content or category
            entries.append({
                'category': category or f'Item {len(entries) + 1}',
                'value': value if value is not None else height,
                'x': (_parse_float(rect.get('x'), 0) or 0) + width / 2,
            })

    entries.sort(key=lambda item: item.get('x', 0))
    categories = [item['category'] for item in entries]
    values = [float(item['value']) for item in entries]
    return {
        'chart_type': 'bar',
        'categories': categories,
        'series': [{'name': 'Series 1', 'values': values}],
        'data_point_count': len(values),
        'extractable': len(values) >= 2,
    }


def _axis_labels(scope: ET.Element, *, vertical: bool) -> list[tuple[float, float]]:
    labels: list[tuple[float, float]] = []
    for text_elem in scope.iter():
        if _local_tag(text_elem.tag) != 'text':
            continue
        content = _element_text(text_elem)
        numeric = _parse_numeric(content)
        if numeric is None:
            continue
        x = _parse_float(text_elem.get('x'), 0) or 0
        y = _parse_float(text_elem.get('y'), 0) or 0
        anchor = (text_elem.get('text-anchor') or '').lower()
        if vertical and anchor == 'end':
            labels.append((y, numeric))
        elif not vertical and anchor == 'middle':
            labels.append((x, numeric if numeric is not None else 0))
    if vertical:
        labels.sort(key=lambda item: item[0], reverse=True)
    else:
        labels.sort(key=lambda item: item[0])
    return labels


def _map_y_to_value(y: float, plot_area: dict[str, Any], y_labels: list[tuple[float, float]]) -> float | None:
    if len(y_labels) >= 2:
        ys = [item[0] for item in y_labels]
        vals = [item[1] for item in y_labels]
        if y <= min(ys):
            return vals[ys.index(min(ys))]
        if y >= max(ys):
            return vals[ys.index(max(ys))]
        for idx in range(len(ys) - 1):
            y_top, y_bottom = ys[idx], ys[idx + 1]
            if y_bottom <= y <= y_top:
                v_top, v_bottom = vals[idx], vals[idx + 1]
                ratio = (y - y_bottom) / (y_top - y_bottom) if y_top != y_bottom else 0
                return v_bottom + ratio * (v_top - v_bottom)
    y_min = plot_area.get('y_min')
    y_max = plot_area.get('y_max')
    if y_min is None or y_max is None or y_max == y_min:
        return None
    ratio = (y - y_min) / (y_max - y_min)
    return max(0.0, 1.0 - ratio) * 100.0


def extract_line_data(root: ET.Element, svg_text: str, chart_root: ET.Element | None = None) -> dict[str, Any]:
    scope = chart_root if chart_root is not None else root
    plot_area = parse_plot_area_marker(svg_text)
    y_labels = _axis_labels(scope, vertical=True)

    x_categories: list[tuple[float, str]] = []
    for text_elem in scope.iter():
        if _local_tag(text_elem.tag) != 'text':
            continue
        if (text_elem.get('text-anchor') or '').lower() != 'middle':
            continue
        label = _element_text(text_elem)
        if not label or _parse_numeric(label) is not None:
            continue
        x = _parse_float(text_elem.get('x'), 0) or 0
        y = _parse_float(text_elem.get('y'), 0) or 0
        if plot_area.get('y_max') and y < plot_area['y_max']:
            continue
        x_categories.append((x, label))
    x_categories.sort(key=lambda item: item[0])

    polylines = scope.findall(f".//{{{SVG_NS}}}polyline")
    if not polylines:
        return {'chart_type': 'line', 'categories': [], 'series': [], 'extractable': False, 'data_point_count': 0}

    polyline = polylines[0]
    points_raw = (polyline.get('points') or '').replace(',', ' ').split()
    coords: list[tuple[float, float]] = []
    for idx in range(0, len(points_raw) - 1, 2):
        try:
            coords.append((float(points_raw[idx]), float(points_raw[idx + 1])))
        except ValueError:
            continue

    if x_categories:
        categories = [label for _, label in x_categories]
        values: list[float] = []
        for px, py in coords:
            value = _map_y_to_value(py, plot_area, y_labels)
            values.append(value if value is not None else 0.0)
        while len(values) < len(categories):
            values.append(0.0)
        values = values[:len(categories)]
    else:
        categories = [f'P{i + 1}' for i in range(len(coords))]
        values = [_map_y_to_value(py, plot_area, y_labels) or 0.0 for _, py in coords]

    return {
        'chart_type': 'line',
        'categories': categories,
        'series': [{'name': 'Series 1', 'values': values}],
        'data_point_count': len(values),
        'extractable': len(values) >= 2,
    }


def extract_pie_data(root: ET.Element, svg_text: str, chart_root: ET.Element | None = None) -> dict[str, Any]:
    scope = chart_root if chart_root is not None else root
    categories: list[str] = []
    values: list[float] = []

    legend = root.find(f".//{{{SVG_NS}}}g[@id='legend']")
    if legend is None:
        legend = scope.find(f".//{{{SVG_NS}}}g[@id='legend']")
    search_roots = [legend] if legend is not None else [scope, root]

    for search_root in search_roots:
        if search_root is None:
            continue
        for text_elem in search_root.iter():
            if _local_tag(text_elem.tag) != 'text':
                continue
            content = _element_text(text_elem)
            if not content:
                continue
            pct_match = PERCENT_RE.search(content)
            if pct_match and '(' in content:
                pct = float(pct_match.group(1))
                label_part = content.split('(')[0].strip()
                if label_part.startswith('$'):
                    continue
                categories.append(label_part or f'Segment {len(categories) + 1}')
                values.append(pct)
                continue
            if pct_match and PERCENT_RE.fullmatch(content.strip()):
                values.append(float(pct_match.group(1)))
                categories.append(f'Segment {len(categories) + 1}')

    if len(values) < 2 and legend is not None:
        rows: list[tuple[str, float]] = []
        pending_label: str | None = None
        for text_elem in legend.iter():
            if _local_tag(text_elem.tag) != 'text':
                continue
            content = _element_text(text_elem)
            if not content or content in ('Market Share', 'Value (Mil)', 'Total Market Size'):
                continue
            pct_match = PERCENT_RE.search(content)
            if pct_match and '(' in content:
                rows.append((pending_label or f'Segment {len(rows) + 1}', float(pct_match.group(1))))
                pending_label = None
            elif _parse_numeric(content) is None and '$' not in content and len(content) < 40:
                pending_label = content
        if len(rows) >= 2:
            categories = [label for label, _ in rows]
            values = [value for _, value in rows]

    return {
        'chart_type': 'pie',
        'categories': categories,
        'series': [{'name': 'Share', 'values': values}],
        'data_point_count': len(values),
        'extractable': len(values) >= 2,
    }


def extract_chart_data(
    svg_path: str | Path,
    chart_type: str = 'auto',
    chart_index: int = 0,
) -> dict[str, Any]:
    """Extract chart data summary from an SVG file."""
    path = Path(svg_path)
    svg_text = path.read_text(encoding='utf-8')
    root = ET.fromstring(svg_text)
    chart_roots = _find_chart_roots(root)
    if not chart_roots:
        return {
            'chart_type': chart_type,
            'categories': [],
            'series': [],
            'extractable': False,
            'data_point_count': 0,
            'error': 'No chart elements found (id containing "chart" or class="chart")',
        }

    if chart_index >= len(chart_roots):
        chart_index = 0
    chart_root = chart_roots[chart_index]
    plot_area = parse_plot_area_marker(svg_text)

    resolved_type = chart_type
    if resolved_type == 'auto':
        resolved_type = detect_chart_type(root, svg_text, chart_root)

    if resolved_type == 'line':
        summary = extract_line_data(root, svg_text, chart_root)
    elif resolved_type == 'pie':
        summary = extract_pie_data(root, svg_text, chart_root)
    else:
        summary = extract_bar_data(root, chart_root)

    summary['plot_area'] = plot_area
    summary['chart_index'] = chart_index
    summary['chart_element_id'] = chart_root.get('id')
    summary['chart_count'] = len(chart_roots)
    if 'error' not in summary and not summary.get('extractable'):
        summary['error'] = 'Insufficient extractable data points'
    return summary


def is_data_heavy_chart(summary: dict[str, Any]) -> bool:
    """Hybrid mode heuristic — prefer Excel charts for dense numeric series."""
    count = int(summary.get('data_point_count') or 0)
    chart_type = summary.get('chart_type')
    if not summary.get('extractable'):
        return False
    if chart_type == 'bar':
        return count >= 5
    if chart_type == 'line':
        return count >= 8
    if chart_type == 'pie':
        return count >= 4
    return count >= 6


def _write_openpyxl_chart(
    summary: dict[str, Any],
    output_path: Path,
) -> None:
    if Workbook is None or BarChart is None:
        raise RuntimeError('openpyxl is required for Excel chart export (pip install openpyxl)')

    wb = Workbook()
    ws = wb.active
    ws.title = 'ChartData'

    categories = summary.get('categories') or []
    series_list = summary.get('series') or []
    if not categories or not series_list:
        raise ValueError('No chart data to write')

    ws.cell(row=1, column=1, value='Category')
    for series_idx, series in enumerate(series_list, start=2):
        ws.cell(row=1, column=series_idx, value=series.get('name') or f'Series {series_idx - 1}')

    for row_idx, category in enumerate(categories, start=2):
        ws.cell(row=row_idx, column=1, value=category)
        for series_idx, series in enumerate(series_list, start=2):
            values = series.get('values') or []
            value = values[row_idx - 2] if row_idx - 2 < len(values) else None
            ws.cell(row=row_idx, column=series_idx, value=value)

    row_count = len(categories) + 1
    chart_type = summary.get('chart_type', 'bar')

    if chart_type == 'line':
        chart = LineChart()
        chart.title = 'Extracted Line Chart'
        chart.y_axis.title = 'Value'
        chart.x_axis.title = 'Category'
        data = Reference(ws, min_col=2, max_col=1 + len(series_list), min_row=1, max_row=row_count)
        cats = Reference(ws, min_col=1, min_row=2, max_row=row_count)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
    elif chart_type == 'pie':
        chart = PieChart()
        chart.title = 'Extracted Pie Chart'
        data = Reference(ws, min_col=2, min_row=1, max_row=row_count)
        cats = Reference(ws, min_col=1, min_row=2, max_row=row_count)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
    else:
        chart = BarChart()
        chart.type = 'col'
        chart.title = 'Extracted Bar Chart'
        chart.y_axis.title = 'Value'
        chart.x_axis.title = 'Category'
        data = Reference(ws, min_col=2, max_col=1 + len(series_list), min_row=1, max_row=row_count)
        cats = Reference(ws, min_col=1, min_row=2, max_row=row_count)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

    ws.add_chart(chart, 'E2')
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))


def svg_chart_to_excel(
    svg_path: str | Path,
    output_xlsx_path: str | Path,
    chart_type: str = 'auto',
    chart_index: int = 0,
) -> dict[str, Any]:
    """Extract SVG chart data and write an XLSX file with a native Excel chart."""
    summary = extract_chart_data(svg_path, chart_type=chart_type, chart_index=chart_index)
    if not summary.get('extractable'):
        raise ValueError(summary.get('error') or 'Chart data is not extractable')

    output_path = Path(output_xlsx_path)
    _write_openpyxl_chart(summary, output_path)
    summary['output_xlsx'] = str(output_path)
    return summary


def list_charts_in_svg(svg_path: str | Path) -> list[dict[str, Any]]:
    """Return extraction summaries for every chart root in an SVG."""
    path = Path(svg_path)
    svg_text = path.read_text(encoding='utf-8')
    root = ET.fromstring(svg_text)
    chart_roots = _find_chart_roots(root)
    summaries: list[dict[str, Any]] = []
    for idx, chart_root in enumerate(chart_roots):
        chart_type = detect_chart_type(root, svg_text, chart_root)
        if chart_type == 'line':
            summary = extract_line_data(root, svg_text, chart_root)
        elif chart_type == 'pie':
            summary = extract_pie_data(root, svg_text, chart_root)
        else:
            summary = extract_bar_data(root, chart_root)
        summary['plot_area'] = parse_plot_area_marker(svg_text)
        summary['chart_index'] = idx
        summary['chart_element_id'] = chart_root.get('id')
        summaries.append(summary)
    return summaries


def verify_chart_data_extractable(svg_path: str | Path) -> tuple[bool, str, dict[str, Any] | None]:
    """Quality-check helper — returns (ok, message, first_chart_summary)."""
    try:
        summaries = list_charts_in_svg(svg_path)
    except ET.ParseError as exc:
        return False, f'Invalid SVG XML: {exc}', None
    except OSError as exc:
        return False, f'Cannot read SVG: {exc}', None

    if not summaries:
        return True, 'No chart elements detected', None

    failed = [item for item in summaries if not item.get('extractable')]
    if failed:
        first = failed[0]
        element_id = first.get('chart_element_id') or 'chart'
        return False, (
            f'Chart data not extractable for #{element_id} '
            f'({first.get("error") or "insufficient data points"})'
        ), first

    first = summaries[0]
    count = first.get('data_point_count', 0)
    ctype = first.get('chart_type', 'unknown')
    return True, f'Extracted {ctype} chart with {count} data point(s)', first


def _cmd_convert(args: argparse.Namespace) -> int:
    try:
        summary = svg_chart_to_excel(args.svg, args.output, chart_type=args.chart_type)
    except Exception as exc:
        print(f'Error: {exc}', file=sys.stderr)
        return 1
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    print(f'Wrote Excel chart: {args.output}')
    return 0


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description='Extract SVG chart data and generate Excel chart workbooks.',
    )
    sub = parser.add_subparsers(dest='command', required=True)

    convert = sub.add_parser('convert', help='Extract chart data and write XLSX')
    convert.add_argument('--svg', required=True, help='Input SVG path')
    convert.add_argument('--output', required=True, help='Output XLSX path')
    convert.add_argument(
        '--chart-type',
        choices=['auto', 'bar', 'line', 'pie'],
        default='auto',
        help='Chart type (default: auto-detect)',
    )
    convert.set_defaults(func=_cmd_convert)
    return parser


def main() -> None:
    parser = build_arg_parser()
    args = parser.parse_args()
    sys.exit(args.func(args))


if __name__ == '__main__':
    main()
